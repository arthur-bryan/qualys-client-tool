from openpyxl import Workbook


class FileManager:
    """
    Class that represents a File Manager.

    Attributes:
        cache (list): list containing dicts with the detections tags and its contents, eg: [{"tag": [vuln1, vuln2]}].
        workbook (:obj:`openpyxl.Workbook`): workbook to save the output data.
        workbook_name: name of the workbook file when saved.
    """

    def __init__(self):
        """Contructor of class FileManager."""
        self.cache = []
        self.workbook = Workbook()
        self.workbook_name = ""

    def add_to_cache(self, tag: str, data: list):
        """
        Adds to the chache, a dict containing the detection tag and the detection, eg: [{"tag": [vuln1, vuln2]}].

        Args:
            tag: title of the sheet, used to identify the detection data.
            data: list containing Asset objects, results from detections.
        """
        self.cache.append({tag: data})

    def get_data_on_cache(self) -> list:
        """
        Returns the entire cache.

        Returns:
            cache: content of the cache.
        """
        return self.cache

    def save_workbook(self, filename: str):
        """
        Iterates over the data on cache, then writes it to a XLSX workbook.

        Args:
            filename: name of the workbook file when its saved.
        """
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        header = ["QID", "SEVERITY", "HOSTNAME", "IP ADDRESS", "STATUS", "TITLE", "FIRST DETECTED", "LAST DETECTED"]
        for index, data in enumerate(self.cache):
            self.workbook.create_sheet(title=list(data.keys())[0])
            sheet = self.workbook.worksheets[index]
            sheet.append(header)
            for asset in list(data.values())[0]:
                for vuln in asset.vulnerabilities:
                    sheet.append([vuln.qid, vuln.severity, asset.hostname, asset.ip, vuln.status,
                                 vuln.title, vuln.first_detected, vuln.last_detected])
        self.workbook.save(filename)
